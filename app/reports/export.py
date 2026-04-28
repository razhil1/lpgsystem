"""
app/reports/export.py
─────────────────────
Excel / XLSX export routes for every report type.
Uses openpyxl for rich, styled spreadsheets — no extra dependencies
beyond what's already in requirements.txt.
"""

from flask import Blueprint, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from app.models import (
    Transaction, TransactionItem, Tank, TankHistory,
    Consumer, Plant, Payment, AuditLog, User
)
from app.extensions import db
from datetime import datetime, date, timedelta
import io
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

export_bp = Blueprint('export', __name__)


# ──────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────

BRAND_COLOR   = "1E3A5F"   # dark navy header
ACCENT_COLOR  = "2563EB"   # blue accent
EVEN_ROW      = "F0F4FF"   # light blue-grey for alternating rows
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color=BRAND_COLOR)
SUB_FONT      = Font(name="Calibri", italic=True, size=10, color="555555")
HEADER_FILL   = PatternFill("solid", fgColor=BRAND_COLOR)
EVEN_FILL     = PatternFill("solid", fgColor=EVEN_ROW)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",  vertical="center", wrap_text=True)
RIGHT         = Alignment(horizontal="right", vertical="center")
THIN          = Side(border_style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT     = '#,##0.00'
DATE_FMT      = 'YYYY-MM-DD'
DATETIME_FMT  = 'YYYY-MM-DD HH:MM'


def _parse_date(val):
    """Safely convert YYYY-MM-DD string → date object."""
    if not val:
        return None
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _period_dates(period: str):
    """
    Return (date_from, date_to) for a named period, or
    fall back to request.args date_from / date_to for 'custom'.
    """
    today = date.today()
    if period == 'daily':
        return today, today
    elif period == 'weekly':
        start = today - timedelta(days=today.weekday())   # Monday
        return start, today
    elif period == 'monthly':
        return today.replace(day=1), today
    elif period == 'yearly':
        return today.replace(month=1, day=1), today
    else:  # custom
        df = _parse_date(request.args.get('date_from')) or today.replace(day=1)
        dt = _parse_date(request.args.get('date_to'))   or today
        return df, dt


def _new_wb(title: str):
    """Create a fresh workbook and return (wb, ws, next_row_ref)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]   # Excel sheet name limit
    return wb, ws


def _write_header_block(ws, title: str, subtitle: str, date_from, date_to, columns):
    """
    Write company header, report title, date range, then column headers.
    Returns the row index where data should start.
    """
    # Row 1 – company
    ws.merge_cells('A1:' + get_column_letter(len(columns)) + '1')
    c = ws['A1']
    c.value = "Gas Pinas Inc. — LPG Management System"
    c.font  = Font(name="Calibri", bold=True, size=13, color=BRAND_COLOR)
    c.alignment = CENTER

    # Row 2 – report title
    ws.merge_cells('A2:' + get_column_letter(len(columns)) + '2')
    c = ws['A2']
    c.value = title
    c.font  = TITLE_FONT
    c.alignment = CENTER

    # Row 3 – subtitle / date range
    ws.merge_cells('A3:' + get_column_letter(len(columns)) + '3')
    c = ws['A3']
    c.value = f"{subtitle}   |   Period: {date_from} to {date_to}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font  = SUB_FONT
    c.alignment = CENTER

    # Row 4 – blank spacer
    ws.row_dimensions[4].height = 6

    # Row 5 – column headers
    for col_idx, (col_label, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=5, column=col_idx, value=col_label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[5].height = 22
    ws.freeze_panes = 'A6'
    return 6   # first data row


def _style_row(ws, row: int, num_cols: int, is_even: bool):
    fill = EVEN_FILL if is_even else None
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = DATA_FONT
        cell.border = BORDER
        if fill:
            cell.fill = fill


def _send_xlsx(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _safe_float(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ──────────────────────────────────────────────────────────
#  EXPORT CENTER (landing page)
# ──────────────────────────────────────────────────────────
@export_bp.route('/')
@login_required
def index():
    from flask import render_template
    return render_template('reports/export_center.html')


# ──────────────────────────────────────────────────────────
#  1. SALES REPORT  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/sales')
@login_required
def sales():
    period = request.args.get('period', 'monthly')
    date_from, date_to = _period_dates(period)

    txns = Transaction.query.filter(
        Transaction.transaction_type == 'Delivery',
        Transaction.transaction_date >= date_from,
        Transaction.transaction_date <= date_to
    ).order_by(Transaction.transaction_date.desc()).all()

    columns = [
        ("Invoice No",       18),
        ("Date",             13),
        ("Consumer",         30),
        ("Tank Sizes / Qty", 22),
        ("Total Amount",     16),
        ("Amount Paid",      16),
        ("Balance",          14),
        ("Payment Status",   16),
        ("Driver",           18),
        ("Truck Plate",      14),
        ("Remarks",          25),
    ]

    wb, ws = _new_wb("Sales Report")
    row = _write_header_block(ws, "Sales Report (Deliveries)", period.title(),
                              date_from, date_to, columns)

    total_sales = total_paid = total_bal = 0.0
    for i, t in enumerate(txns):
        sizes_qty = ", ".join(f"{s}×{q}" for s, q in t.qty_by_size.items())
        amt   = _safe_float(t.total_amount)
        paid  = _safe_float(t.amount_paid)
        bal   = amt - paid
        total_sales += amt
        total_paid  += paid
        total_bal   += bal

        row_data = [
            t.invoice_no,
            t.transaction_date,
            t.consumer.business_name if t.consumer else "—",
            sizes_qty,
            amt,
            paid,
            bal,
            t.payment_status,
            t.driver_name or "—",
            t.truck_plate or "—",
            t.remarks or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, date) and not isinstance(val, datetime):
                cell.number_format = DATE_FMT
                cell.alignment = CENTER
            elif col_idx in (5, 6, 7):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="TOTALS:").font = Font(name="Calibri", bold=True, size=10)
    for col, val in [(5, total_sales), (6, total_paid), (7, total_bal)]:
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = MONEY_FMT
        c.font = Font(name="Calibri", bold=True, size=10, color=BRAND_COLOR)
        c.alignment = RIGHT

    fname = f"Sales_Report_{period}_{date_from}_to_{date_to}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  2. TRANSACTIONS REPORT (all types)  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/transactions')
@login_required
def transactions():
    period   = request.args.get('period', 'monthly')
    txn_type = request.args.get('type', '')
    date_from, date_to = _period_dates(period)

    q = Transaction.query.filter(
        Transaction.transaction_date >= date_from,
        Transaction.transaction_date <= date_to
    )
    if txn_type:
        q = q.filter(Transaction.transaction_type == txn_type)
    txns = q.order_by(Transaction.transaction_date.desc()).all()

    columns = [
        ("Invoice No",     18),
        ("Type",           14),
        ("Date",           13),
        ("Consumer/Plant", 28),
        ("Qty",             8),
        ("Total Amount",   16),
        ("Paid",           14),
        ("Balance",        13),
        ("Status",         14),
        ("Created By",     18),
        ("Created At",     18),
    ]

    wb, ws = _new_wb("All Transactions")
    subtitle = f"{txn_type or 'All Types'}"
    row = _write_header_block(ws, "Transactions Report", subtitle,
                              date_from, date_to, columns)

    for i, t in enumerate(txns):
        party = "—"
        if t.consumer:
            party = t.consumer.business_name
        elif t.plant:
            party = t.plant.plant_name

        amt  = _safe_float(t.total_amount)
        paid = _safe_float(t.amount_paid)
        bal  = amt - paid

        row_data = [
            t.invoice_no,
            t.transaction_type,
            t.transaction_date,
            party,
            t.total_qty,
            amt,
            paid,
            bal,
            t.payment_status,
            t.creator.full_name if t.creator else "—",
            t.created_at,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment = CENTER
            elif isinstance(val, date):
                cell.number_format = DATE_FMT
                cell.alignment = CENTER
            elif col_idx in (6, 7, 8):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    fname = f"Transactions_{subtitle.replace(' ','_')}_{date_from}_to_{date_to}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  3. TANK INVENTORY  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/tanks')
@login_required
def tanks():
    status   = request.args.get('status', '')
    location = request.args.get('location', '')

    q = Tank.query.filter_by(is_active=True)
    if status:
        q = q.filter(Tank.status == status)
    if location:
        q = q.filter(Tank.location == location)
    all_tanks = q.order_by(Tank.tank_size, Tank.serial_number).all()

    columns = [
        ("Serial / Code",   20),
        ("Has Serial",      12),
        ("Batch Code",      20),
        ("Size",            10),
        ("Brand",           14),
        ("Status",          14),
        ("Location",        16),
        ("Consumer",        26),
        ("Plant",           22),
        ("Purchase Date",   15),
        ("Purchase Cost",   15),
        ("Last Txn Date",   18),
        ("Notes",           25),
    ]

    label_parts = []
    if status:
        label_parts.append(f"Status={status}")
    if location:
        label_parts.append(f"Location={location}")
    subtitle = ", ".join(label_parts) if label_parts else "All Active Tanks"

    wb, ws = _new_wb("Tank Inventory")
    today = date.today()
    row = _write_header_block(ws, "Tank Inventory Export", subtitle,
                              today, today, columns)

    for i, t in enumerate(all_tanks):
        row_data = [
            t.serial_number or "—",
            "Yes" if t.has_serial_number else "No",
            t.batch_code or "—",
            t.tank_size,
            t.brand or "—",
            t.status,
            t.location,
            t.current_consumer.business_name if t.current_consumer else "—",
            t.current_plant.plant_name if t.current_plant else "—",
            t.purchase_date,
            _safe_float(t.purchase_cost),
            t.last_transaction_date,
            t.notes or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment = CENTER
            elif isinstance(val, date):
                cell.number_format = DATE_FMT
                cell.alignment = CENTER
            elif col_idx == 11:
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    fname = f"Tank_Inventory_{today}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  4. OUTSTANDING BALANCES  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/outstanding')
@login_required
def outstanding():
    txns = Transaction.query.filter(
        Transaction.payment_status.in_(['Unpaid', 'Partial']),
        Transaction.transaction_type == 'Delivery'
    ).order_by(Transaction.transaction_date).all()

    columns = [
        ("Invoice No",      18),
        ("Date",            13),
        ("Consumer",        30),
        ("Total Amount",    16),
        ("Amount Paid",     16),
        ("Balance Due",     15),
        ("Payment Status",  16),
        ("Days Outstanding",18),
    ]

    wb, ws = _new_wb("Outstanding Balances")
    today = date.today()
    row = _write_header_block(ws, "Outstanding Balances Report",
                              "Unpaid & Partial Deliveries",
                              today, today, columns)

    grand_total = grand_paid = grand_bal = 0.0
    for i, t in enumerate(txns):
        amt  = _safe_float(t.total_amount)
        paid = _safe_float(t.amount_paid)
        bal  = amt - paid
        grand_total += amt
        grand_paid  += paid
        grand_bal   += bal

        row_data = [
            t.invoice_no,
            t.transaction_date,
            t.consumer.business_name if t.consumer else "—",
            amt,
            paid,
            bal,
            t.payment_status,
            t.days_outstanding,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, date) and not isinstance(val, datetime):
                cell.number_format = DATE_FMT
                cell.alignment = CENTER
            elif col_idx in (4, 5, 6):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            elif col_idx == 8:
                cell.alignment = CENTER
                # Highlight overdue rows (> 30 days)
                if isinstance(val, int) and val > 30:
                    cell.font = Font(name="Calibri", size=10, color="C0392B", bold=True)
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    # Grand totals
    row += 1
    ws.cell(row=row, column=3, value="GRAND TOTAL:").font = Font(name="Calibri", bold=True)
    for col, val in [(4, grand_total), (5, grand_paid), (6, grand_bal)]:
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = MONEY_FMT
        c.font = Font(name="Calibri", bold=True, size=10, color=BRAND_COLOR)
        c.alignment = RIGHT

    fname = f"Outstanding_Balances_{today}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  5. TANK MOVEMENT HISTORY  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/tank-movement')
@login_required
def tank_movement():
    period = request.args.get('period', 'monthly')
    date_from, date_to = _period_dates(period)
    date_to_dt = datetime.combine(date_to, datetime.max.time())

    history = TankHistory.query.filter(
        TankHistory.created_at >= datetime.combine(date_from, datetime.min.time()),
        TankHistory.created_at <= date_to_dt
    ).order_by(TankHistory.created_at.desc()).all()

    columns = [
        ("Tank Code",        20),
        ("Tank Size",        12),
        ("Event",            18),
        ("Description",      35),
        ("From",             18),
        ("To",               18),
        ("Consumer",         26),
        ("Plant",            22),
        ("Date & Time",      20),
        ("Done By",          18),
    ]

    wb, ws = _new_wb("Tank Movement")
    row = _write_header_block(ws, "Tank Movement History", period.title(),
                              date_from, date_to, columns)

    for i, h in enumerate(history):
        row_data = [
            h.tank.serial_number if h.tank else "—",
            h.tank.tank_size if h.tank else "—",
            h.event_type,
            h.event_description or "",
            h.from_location or "—",
            h.to_location or "—",
            h.consumer.business_name if h.consumer else "—",
            h.plant.plant_name if h.plant else "—",
            h.created_at,
            h.creator.full_name if h.creator else "—",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    fname = f"Tank_Movement_{period}_{date_from}_to_{date_to}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  6. CONSUMERS LIST  →  Excel
# ──────────────────────────────────────────────────────────
@export_bp.route('/consumers')
@login_required
def consumers():
    all_consumers = Consumer.query.filter_by(is_active=True)\
                                  .order_by(Consumer.business_name).all()

    columns = [
        ("Code",              14),
        ("Business Name",     30),
        ("Type",              16),
        ("Contact Person",    22),
        ("Phone",             16),
        ("Email",             26),
        ("Address",           30),
        ("Credit Limit",      15),
        ("Outstanding Bal",   17),
        ("Registered On",     15),
    ]

    today = date.today()
    wb, ws = _new_wb("Consumers")
    row = _write_header_block(ws, "Consumer Masterlist", "All Active Consumers",
                              today, today, columns)

    for i, c in enumerate(all_consumers):
        row_data = [
            c.consumer_code,
            c.business_name,
            c.consumer_type,
            c.contact_person or "—",
            c.phone or "—",
            c.email or "—",
            c.address or "—",
            _safe_float(c.credit_limit),
            _safe_float(c.outstanding_balance),
            c.created_at,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment = CENTER
            elif col_idx in (8, 9):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    fname = f"Consumers_{today}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  7. AUDIT TRAIL  →  Excel  (admin only)
# ──────────────────────────────────────────────────────────
@export_bp.route('/audit-trail')
@login_required
def audit_trail():
    if current_user.role != 'admin':
        flash('Access denied. Admins only.', 'danger')
        return redirect(url_for('export.index'))

    period = request.args.get('period', 'monthly')
    date_from, date_to = _period_dates(period)
    date_to_dt = datetime.combine(date_to, datetime.max.time())

    logs = AuditLog.query.filter(
        AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()),
        AuditLog.created_at <= date_to_dt
    ).order_by(AuditLog.created_at.desc()).all()

    columns = [
        ("Date & Time",  20),
        ("User",         20),
        ("Action",       12),
        ("Module",       14),
        ("Record ID",    12),
        ("Description",  45),
        ("IP Address",   16),
    ]

    wb, ws = _new_wb("Audit Trail")
    row = _write_header_block(ws, "Audit Trail", period.title(),
                              date_from, date_to, columns)

    action_colors = {
        'CREATE': '27AE60', 'UPDATE': '2980B9',
        'DELETE': 'C0392B', 'LOGIN':  '8E44AD',
    }

    for i, log in enumerate(logs):
        row_data = [
            log.created_at,
            log.user.full_name if log.user else "—",
            log.action,
            log.module,
            log.record_id or "—",
            log.description or "",
            log.ip_address or "—",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment = CENTER
            elif col_idx == 3:   # Action column — colour-coded
                cell.alignment = CENTER
                color = action_colors.get(str(val), '555555')
                cell.font = Font(name="Calibri", size=10, bold=True, color=color)
            else:
                cell.alignment = LEFT
        _style_row(ws, row, len(columns), i % 2 == 1)
        row += 1

    fname = f"Audit_Trail_{period}_{date_from}_to_{date_to}.xlsx"
    return _send_xlsx(wb, fname)


# ──────────────────────────────────────────────────────────
#  8. SUMMARY DASHBOARD EXPORT  →  Excel (single sheet overview)
# ──────────────────────────────────────────────────────────
@export_bp.route('/summary')
@login_required
def summary():
    period = request.args.get('period', 'monthly')
    date_from, date_to = _period_dates(period)
    today = date.today()

    # ── gather data ─────────────────────────────────────
    deliveries = Transaction.query.filter(
        Transaction.transaction_type == 'Delivery',
        Transaction.transaction_date >= date_from,
        Transaction.transaction_date <= date_to
    ).all()

    total_sales     = sum(_safe_float(t.total_amount) for t in deliveries)
    total_paid      = sum(_safe_float(t.amount_paid)  for t in deliveries)
    total_bal       = total_sales - total_paid

    tank_counts = {}
    for size in ['11kg', '11kg Fiber', '22kg', '50kg', 'Industrial']:
        tank_counts[size] = {
            'total':     Tank.query.filter_by(tank_size=size, is_active=True).count(),
            'full':      Tank.query.filter_by(tank_size=size, status='Full', is_active=True).count(),
            'empty':     Tank.query.filter_by(tank_size=size, status='Empty', is_active=True).count(),
            'at_plant':  Tank.query.filter_by(tank_size=size, location='At Plant', is_active=True).count(),
            'with_cons': Tank.query.filter_by(tank_size=size, location='With Consumer', is_active=True).count(),
        }

    unpaid_count = Transaction.query.filter(
        Transaction.payment_status.in_(['Unpaid', 'Partial']),
        Transaction.transaction_type == 'Delivery'
    ).count()

    wb, ws = _new_wb("Summary")
    ws.sheet_view.showGridLines = False

    # ── Title block ─────────────────────────────────────
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "Gas Pinas Inc. — LPG Management System"
    c.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill  = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f"SUMMARY REPORT  |  {period.title()}  |  {date_from} to {date_to}"
    c.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=ACCENT_COLOR)
    c.alignment = CENTER
    ws.row_dimensions[2].height = 22

    ws.merge_cells('A3:F3')
    c = ws['A3']
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font  = SUB_FONT
    c.alignment = CENTER

    def _section_header(row, label):
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = label
        c.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="374151")
        c.alignment = LEFT
        ws.row_dimensions[row].height = 20

    def _kv(row, label, value, fmt=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name="Calibri", size=10, bold=True)
        lc.alignment = LEFT
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = RIGHT
        if fmt:
            vc.number_format = fmt

    # ── Sales Summary ────────────────────────────────────
    r = 5
    _section_header(r, "  📊 Sales Summary")
    r += 1
    for label, val, fmt in [
        ("Total Deliveries",  len(deliveries),  None),
        ("Total Sales Amount", total_sales,      MONEY_FMT),
        ("Total Collected",    total_paid,       MONEY_FMT),
        ("Outstanding Balance",total_bal,        MONEY_FMT),
        ("Unpaid Invoices",    unpaid_count,     None),
    ]:
        _kv(r, label, val, fmt)
        r += 1

    # ── Tank Inventory ───────────────────────────────────
    r += 1
    _section_header(r, "  🪣 Tank Inventory by Size")
    r += 1
    # header row
    for col, label, width in [
        (1,"Size",12),(2,"Total",10),(3,"Full",10),
        (4,"Empty",10),(5,"At Plant",12),(6,"With Consumer",16)
    ]:
        c = ws.cell(row=r, column=col, value=label)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    r += 1

    for ii, (size, counts) in enumerate(tank_counts.items()):
        for col, val in enumerate([
            size,
            counts['total'], counts['full'],
            counts['empty'], counts['at_plant'], counts['with_cons']
        ], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = CENTER if col > 1 else LEFT
            cell.border = BORDER
            if ii % 2 == 1:
                cell.fill = EVEN_FILL
        r += 1

    fname = f"Summary_Report_{period}_{date_from}_to_{date_to}.xlsx"
    return _send_xlsx(wb, fname)
